import os

from pysat.solvers import Glucose3
from sat_solver.sat_algorithms.utils import VariableFactory
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Equation (6): Ensure each task has one unique start time
def encode_unique_start(solver, vf, max_time, task_id, duration):
    start_vars = [vf.start(task_id, t) for t in range(max_time - duration + 1)]
    solver.add_clause(start_vars)
    for i in range(len(start_vars)):
        for j in range(i + 1, len(start_vars)):
            solver.add_clause([-start_vars[i], -start_vars[j]])

# Equation (7): Mã hóa ràng buộc giới hạn thời gian bắt đầu
def encode_start_in_time(solver, vf, max_time, task_id, duration):
    """
    Mã hóa ràng buộc đảm bảo rằng hoạt động không được bắt đầu sau thời điểm kết thúc có thể.
    - Tạo các mệnh đề phủ định để loại bỏ các thời điểm không hợp lệ.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - task_id: ID của hoạt động.
    - duration: Thời gian thực hiện hoạt động.
    """
    # Loại bỏ các thời điểm không hợp lệ (bắt đầu sau thời điểm max_time - duration)
    for t in range(max_time - duration + 1, max_time):
        solver.add_clause([-vf.start(task_id, t)])

# Equation (8) và (9): Mã hóa ràng buộc liên tục khi hoạt động đang chạy
def encode_runtime(solver, vf, max_time, task_id, duration):
    """
    Mã hóa ràng buộc đảm bảo rằng:
    - Nếu hoạt động bắt đầu tại thời điểm t, nó phải chạy liên tục trong khoảng từ t đến t + duration - 1.
    - Hoạt động không thể chạy ở thời điểm khác ngoài khoảng thời gian này.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - task_id: ID của công việc.
    - duration: Thời gian thực hiện công việc.
    """
    for t in range(max_time):
        start_var = vf.start(task_id, t)
        for j in range(max_time):
            run_var = vf.run(task_id, j)
            # Nếu ngoài khoảng thời gian thực hiện, không thể chạy
            if j < t or j >= t + duration:
                solver.add_clause([-start_var, -run_var])
            # Nếu trong khoảng thời gian thực hiện, phải chạy
            else:
                solver.add_clause([-start_var, run_var])

# Equation (10): Mã hóa quan hệ "Finish-to-Start"
def encode_relation_fs(solver, vf, max_time, task1, task2, duration1):
    """
    Mã hóa ràng buộc "Finish-to-Start" giữa hai công việc.
    - Nếu công việc 1 kết thúc tại thời điểm t, công việc 2 không thể bắt đầu trước thời điểm t + duration1.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - task1: ID của công việc 1.
    - task2: ID của công việc 2.
    - duration1: Thời gian thực hiện công việc 1.
    """
    for t in range(max_time):
        start_var = vf.start(task1, t)
        # Prevent task2 from starting before task1 finishes
        for k in range(t + duration1):
            solver.add_clause([-start_var, -vf.start(task2, k)])

def encode_resource_constraint_powerset(solver, vf, max_time, tasks, resources):
    """
    Mã hóa ràng buộc tài nguyên bằng tập lũy thừa (Powerset).
    Mục đích: Mã hóa xung đột tài nguyên bằng cách xét tất cả các tập hợp con của hoạt động tại mỗi thời điểm.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - tasks: Danh sách các hoạt động.
    - resources: Danh sách các tài nguyên.
    """
    for t in range(max_time):
        for resource in resources:
            # Get all tasks that use this resource
            consuming_tasks = []
            for task in tasks:
                consumption = abs(task.get("consumption", {}).get(resource["id"], 0))
                if consumption > 0:
                    consuming_tasks.append({"task": task, "consumption": consumption})

            # Generate all possible combinations that would exceed capacity
            from itertools import combinations
            for r in range(1, len(consuming_tasks) + 1):
                for combo in combinations(consuming_tasks, r):
                    total_consumption = sum(item["consumption"] for item in combo)
                    if total_consumption > resource["capacity"]:
                        # Add clause preventing this combination
                        clause = [-vf.run(item["task"]["id"], t) for item in combo]
                        solver.add_clause(clause)

def decode_solution(tasks, model, vf, consumptions):
    """
    Decode the SAT solver's model to extract task start times and schedule details.

    :param tasks: List of task dictionaries
    :param model: Raw model from the SAT solver
    :param vf: VariableFactory used in encoding
    :return: Decoded schedule information
    """
    # Create reverse mapping for variables
    reverse_var_map = {v: k for k, v in vf.var_map.items()}

    # Extract positive variables from the model
    positive_vars = set(abs(var) for var in model if var > 0)

    # Store start times for tasks
    task_start_times = {}

    # Decode start times
    for task in tasks:
        task_start_found = False
        for time in range(len(reverse_var_map)):
            start_var = vf.start(task['id'], time)
            if start_var in positive_vars:
                task_start_times[task['id']] = time
                task_start_found = True
                break

        if not task_start_found:
            print(f"Warning: No start time found for task {task['id']}")

    # Sort tasks by start time
    sorted_tasks = sorted(tasks, key=lambda x: task_start_times.get(x['id'], float('inf')))

    # Prepare detailed schedule
    schedule = []
    for task in sorted_tasks:
        start_time = task_start_times.get(task['id'])
        if start_time is not None:
            end_time = start_time + task['duration']

            # Collect resource consumption
            task_resources = []
            for consumption in consumptions:
                if consumption['task_id'] == task['id']:
                    task_resources.append({
                        'resource_id': consumption['resource_id'],
                        'amount': consumption['amount']
                    })

            schedule.append({
                'task_id': task['id'],
                'task_name': task['name'],
                'start_time': start_time,
                'end_time': end_time,
                'duration': task['duration'],
                'resources_consumed': task_resources
            })

    return schedule

def solve_rcpsp(max_time, tasks, relations, consumptions, resources):
    solver = Glucose3()
    vf = VariableFactory()

    # Encoding tasks with start time, runtime constraints
    for task in tasks:
        encode_unique_start(solver, vf, max_time, task["id"], task["duration"])
        encode_start_in_time(solver, vf, max_time, task["id"], task["duration"])
        encode_runtime(solver, vf, max_time, task["id"], task["duration"])

    # Encoding precedence relations (Finish-to-Start)
    for relation in relations:
        task_1 = relation["task_id_1"]
        task_2 = relation["task_id_2"]
        task_1_duration = next(t["duration"] for t in tasks if t["id"] == task_1)
        encode_relation_fs(solver, vf, max_time, task_1, task_2, task_1_duration)

    # Encoding resource consumption
    for task in tasks:
        task.setdefault("consumption", {})

    for consumption in consumptions:
        task = next(t for t in tasks if t["id"] == consumption["task_id"])
        task["consumption"][consumption["resource_id"]] = consumption["amount"]

    # Encoding resource constraints
    encode_resource_constraint_powerset(solver, vf, max_time, tasks, resources)

    # Solve the problem
    if solver.solve():
        model = solver.get_model()
        status = "SAT"
        # Solver is returned along with other parts
        return model, vf, solver, status
    else:
        status = "UNSAT"
        solver.delete()
        return None, None, solver, status

def export_schedule_to_xlsx(
        schedule, solver, vf, status, tasks, resources, relations, start_task_id,
        output_file, append=False):
    """
    Export schedule data to an Excel (.xlsx) file with simplified fields and plain unique sequential task IDs.

    Args:
        schedule (list): List of scheduled tasks.
        solver: SAT solver instance (for getting variable and clause counts).
        vf: VariableFactory instance.
        status (str): Status (e.g., "SAT" or "UNSAT").
        tasks (list): List of tasks.
        resources (list): List of resources.
        relations (list): List of relations.
        start_task_id (int): Starting ID for the task numbering.
        output_file (str): Path to output Excel file.
        append (bool): Whether to append to an existing workbook or create a new one.
    """
    # Calculate problem metrics
    num_variables = solver.nof_vars()
    num_clauses = solver.nof_clauses()
    num_tasks = len(tasks)
    num_resources = len(resources)
    num_relations = len(relations)
    problem_field = f"{num_tasks}-{num_resources}-{num_relations}"

    # Prepare data for Excel export
    excel_data = []
    current_task_id = start_task_id
    for task in schedule:
        # Calculate execution time
        execution_time = task['end_time'] - task['start_time']

        # Append row data
        row = [
            current_task_id,  # Task ID
            problem_field,  # Problem
            'powerset',  # Type
            status,  # Status (SAT/UNSAT)
            execution_time,  # Time
            num_variables,  # Variables
            num_clauses  # Clauses
        ]
        excel_data.append(row)

        # Increment the task ID for the next task
        current_task_id += 1

    # Define column headers
    headers = ['Task ID', 'Problem', 'Type', 'Status', 'Time', 'Variables', 'Clauses']

    # Write to Excel file
    if append and os.path.exists(output_file):
        # Load the existing workbook if appending
        from openpyxl import load_workbook
        workbook = load_workbook(output_file)
        sheet = workbook.active
        starting_row = sheet.max_row + 1  # Start appending after the last row
    else:
        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        starting_row = 1  # Write headers if creating a new file

        # Write headers in the first row
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{starting_row}'] = header
        starting_row += 1  # Move to the next row for data

    # Write data rows to the worksheet
    for row_num, row_data in enumerate(excel_data, starting_row):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data

    # Save the changes to the output file
    workbook.save(output_file)

    # Return the last task ID used
    return current_task_id

