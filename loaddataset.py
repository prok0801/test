import os
from pathlib import Path

from pypblib import pblib
from pypblib.pblib import PBConfig, Pb2cnf
from pysat.solvers import Glucose3
from sat_solver.sat_algorithms.utils import VariableFactory

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ----------------- Các hàm mã hóa ràng buộc -----------------

# Equation (6): Mã hóa ràng buộc hoạt động phải bắt đầu tại một thời điểm duy nhất (ALK - AtLeastK)
def encode_unique_start_instant_alk(solver, vf, max_time, task_id, duration):
    """
    Mã hóa ràng buộc đảm bảo rằng một hoạt động chỉ có thể bắt đầu tại một thời điểm duy nhất.
    - Tạo một mệnh đề (clause) để đảm bảo rằng ít nhất một thời điểm bắt đầu được chọn.
    - Tạo thêm các mệnh đề để đảm bảo rằng không có hai thời điểm bắt đầu được chọn đồng thời.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa có thể bắt đầu hoạt động.
    - task_id: ID của hoạt động.
    - duration: Thời gian thực hiện hoạt động.
    """
    # Tạo mệnh đề để đảm bảo hoạt động phải bắt đầu tại ít nhất một thời điểm
    clause = [vf.start(task_id, t) for t in range(max_time - duration + 1)]
    solver.add_clause(clause)

    # Tạo các mệnh đề để đảm bảo hoạt động chỉ bắt đầu tại duy nhất một thời điểm
    for t1 in range(max_time - duration + 1):
        for t2 in range(t1 + 1, max_time - duration + 1):
            solver.add_clause([-vf.start(task_id, t1), -vf.start(task_id, t2)])

# Equation (7): Mã hóa ràng buộc giới hạn thời gian bắt đầu
def encode_start_in_time(solver, vf, max_time, task_id, duration):
    """
    Mã hóa ràng buộc đảm bảo rằng hoạt động không được bắt đầu sau thời điểm kết thúc có thể.
    - Tạo các mệnh đề phủ định để loại bỏ các thời điểm không hợp lệ.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - task_id: ID của hoạt động.
    - duration: Thời gian thực hiện hoạt động.
    """
    # Loại bỏ các thời điểm không hợp lệ (bắt đầu sau thời điểm max_time - duration)
    for t in range(max_time - duration + 1, max_time):
        solver.add_clause([-vf.start(task_id, t)])

# Equation (8) và (9): Mã hóa ràng buộc liên tục khi hoạt động đang chạy
def encode_runtime(solver, vf, max_time, task_id, duration):
    """
    Mã hóa ràng buộc đảm bảo rằng:
    - Nếu hoạt động bắt đầu tại thời điểm t, nó phải chạy liên tục trong khoảng từ t đến t + duration - 1.
    - Hoạt động không thể chạy ở thời điểm khác ngoài khoảng thời gian này.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - task_id: ID của công việc.
    - duration: Thời gian thực hiện công việc.
    """
    for t in range(max_time):
        start_var = vf.start(task_id, t)
        for j in range(max_time):
            run_var = vf.run(task_id, j)
            # Nếu ngoài khoảng thời gian thực hiện, không thể chạy
            if j < t or j >= t + duration:
                solver.add_clause([-start_var, -run_var])
            # Nếu trong khoảng thời gian thực hiện, phải chạy
            else:
                solver.add_clause([-start_var, run_var])

# Equation (10): Mã hóa quan hệ "Finish-to-Start"
def encode_relation_fs(solver, vf, max_time, task1, task2, duration1):
    """
    Mã hóa ràng buộc "Finish-to-Start" giữa hai công việc.
    - Nếu công việc 1 kết thúc tại thời điểm t, công việc 2 không thể bắt đầu trước thời điểm t + duration1.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - task1: ID của công việc 1.
    - task2: ID của công việc 2.
    - duration1: Thời gian thực hiện công việc 1.
    """
    for t in range(max_time):
        start_var = vf.start(task1, t)
        for k in range(t + duration1):  # Changed this part
            solver.add_clause([-start_var, -vf.start(task2, k)])

# Equation (16): Mã hóa atoms tiêu thụ tài nguyên
def encode_consumption_atoms(solver, vf, max_time, tasks, resources):
    """
    Mã hóa ràng buộc tiêu thụ tài nguyên của các hoạt động:
    - Nếu một công việc đang chạy tại thời điểm t, nó phải tiêu thụ một lượng tài nguyên nhất định.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - tasks: Danh sách các hoạt động.
    - resources: Danh sách các tài nguyên.
    """
    for t in range(max_time):
        for task in tasks:
            for resource in resources:
                # Lấy số lượng tài nguyên tiêu thụ
                consumption = task.get("consumption", {}).get(resource["id"], 0)
                for i in range(consumption):
                    # Nếu công việc đang chạy, tài nguyên phải được tiêu thụ
                    solver.add_clause([-vf.run(task["id"], t), vf.consume(task["id"], resource["id"], t, i)])

# Equation (5 and 17): BCC resource constraint
def encode_resource_constraint_cardinality(solver, vf, max_time, tasks, resources):
    """
    Mã hóa ràng buộc tài nguyên bằng BCC sử dụng Sequential Counter (NSC) với PBLib.

    Parameters:
    - solver: Bộ giải SAT.
    - vf: Lớp quản lý các biến (VariableFactory).
    - max_time: Thời gian tối đa.
    - tasks: Danh sách các hoạt động.
    - resources: Danh sách các tài nguyên.
    """
    id_variable = int  # Biến toàn cục để đếm số lượng biến được tạo ra trong quá trình mã hóa

    pbConfig = PBConfig()  # Cấu hình cho PBLib
    pbConfig.set_AMK_Encoder(pblib.AMK_CARD)  # Dùng NSC Encoder cho AtMostK (NSC là một thuật toán mã hóa cho BCC)
    pb2 = Pb2cnf(pbConfig)  # Tạo đối tượng Pb2cnf để mã hóa các ràng buộc

    # Lặp qua từng thời điểm từ 0 đến max_time
    for t in range(max_time):
        # Lặp qua từng tài nguyên
        for resource in resources:
            # Lấy các biến tiêu thụ tài nguyên cho từng công việc tại thời điểm t
            consumption_vars = [
                vf.consume(task["id"], resource["id"], t, i)  # Biến tiêu thụ tài nguyên của công việc tại thời điểm t
                for task in tasks
                for i in range(task.get("consumption", {}).get(resource["id"], 0))  # Lấy số lượng tài nguyên công việc tiêu thụ
            ]

            # Nếu có các biến tiêu thụ tài nguyên (tức là có công việc tiêu thụ tài nguyên tại thời điểm t)
            if consumption_vars:
                # Khởi tạo công thức cho BCC
                formula = []
                weights = [1] * len(consumption_vars)  # Tạo trọng số cho các biến (tất cả trọng số là 1)

                # Mã hóa ràng buộc "AtMostK" với NSC: đảm bảo rằng không vượt quá dung lượng tài nguyên
                max_var = pb2.encode_at_most_k(
                    weights, consumption_vars, resource["capacity"], formula, id_variable
                )

                # Cập nhật id_variable sau khi mã hóa
                id_variable = max_var + 1

                # Thêm các ràng buộc (clauses) vào solver
                for clause in formula:
                    solver.add_clause(clause)

# ----------------- Hàm giải mã và giải bài toán RCPSP -----------------

def decode_solution(tasks, model, vf, consumptions):
    """
    Decode the SAT solver's model to extract task start times and schedule details.

    :param tasks: List of task dictionaries
    :param model: Raw model from the SAT solver
    :param vf: VariableFactory used in encoding
    :return: Decoded schedule information
    """
    # Create reverse mapping for variables
    reverse_var_map = {v: k for k, v in vf.var_map.items()}

    # Extract positive variables from the model
    positive_vars = set(abs(var) for var in model if var > 0)

    # Store start times for tasks
    task_start_times = {}

    # Decode start times
    for task in tasks:
        task_start_found = False
        for time in range(len(reverse_var_map)):
            start_var = vf.start(task['id'], time)
            if start_var in positive_vars:
                task_start_times[task['id']] = time
                task_start_found = True
                break

        if not task_start_found:
            print(f"Warning: No start time found for task {task['id']}")

    # Sort tasks by start time
    sorted_tasks = sorted(tasks, key=lambda x: task_start_times.get(x['id'], float('inf')))

    # Prepare detailed schedule
    schedule = []
    for task in sorted_tasks:
        start_time = task_start_times.get(task['id'])
        if start_time is not None:
            end_time = start_time + task['duration']

            # Collect resource consumption
            task_resources = []
            for consumption in consumptions:
                if consumption['task_id'] == task['id']:
                    task_resources.append({
                        'resource_id': consumption['resource_id'],
                        'amount': consumption['amount']
                    })

            schedule.append({
                'task_id': task['id'],
                'task_name': task['name'],
                'start_time': start_time,
                'end_time': end_time,
                'duration': task['duration'],
                'resources_consumed': task_resources
            })

    return schedule

def solve_rcpsp(max_time, tasks, relations, consumptions, resources):
    solver = Glucose3()
    vf = VariableFactory()

    # Encoding tasks with ALK, start time, runtime constraints
    for task in tasks:
        encode_unique_start_instant_alk(solver, vf, max_time, task["id"], task["duration"])
        encode_start_in_time(solver, vf, max_time, task["id"], task["duration"])
        encode_runtime(solver, vf, max_time, task["id"], task["duration"])

    # Encoding precedence relations (Finish-to-Start)
    for relation in relations:
        task_1 = relation["task_id_1"]
        task_2 = relation["task_id_2"]
        task_1_duration = next(t["duration"] for t in tasks if t["id"] == task_1)
        encode_relation_fs(solver, vf, max_time, task_1, task_2, task_1_duration)

    # Encoding resource consumption
    for task in tasks:
        task.setdefault("consumption", {})

    for consumption in consumptions:
        task = next(t for t in tasks if t["id"] == consumption["task_id"])
        task["consumption"][consumption["resource_id"]] = consumption["amount"]

    # Encoding resource constraints
    encode_resource_constraint_cardinality(solver, vf, max_time, tasks, resources)

    # Solve the problem and calculate variables & clauses
    variables, clauses = solver.nof_vars(), solver.nof_clauses()

    if solver.solve():
        model = solver.get_model()
        status = "SAT"
        solver.delete()
        return model, vf, variables, clauses, status  # Capture variables and clauses
    else:
        status = "UNSAT"
        solver.delete()
        return None, None, variables, clauses, status

# ----------------- Hàm xuất kết quả ra Excel -----------------

def export_schedule_to_xlsx(
        schedule, variables, clauses, status, tasks, resources, relations, start_task_id,
        output_file, append=False):
    """
    Export schedule data to an Excel (.xlsx) file with consistent formatting and plain unique sequential task IDs.
    """
    headers = ['Task ID', 'Problem', 'Type', 'Status', 'Time', 'Variables', 'Clauses']

    num_tasks = len(tasks)
    num_resources = len(resources)
    num_relations = len(relations)
    problem_field = f"{num_tasks}-{num_resources}-{num_relations}"

    excel_data = []
    current_task_id = start_task_id  

    for task in schedule:
        execution_time = task['end_time'] - task['start_time']

        row = [
            current_task_id,  
            problem_field,     
            'bcc',              
            status,            
            execution_time,     
            variables,          
            clauses             
        ]
        excel_data.append(row)
        current_task_id += 1

    if append and os.path.exists(output_file):
        workbook = load_workbook(output_file)
        sheet = workbook.active
        starting_row = sheet.max_row + 1
    else:
        workbook = Workbook()
        sheet = workbook.active
        starting_row = 1
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}{starting_row}"] = header
        starting_row += 1

    for row_num, row_data in enumerate(excel_data, starting_row):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}{row_num}"] = cell_data

    workbook.save(output_file)
    return current_task_id

# ----------------- Hàm đọc và phân tích file .smt -----------------

def parse_smt_file(file_path):
    """
    Đọc và phân tích file .smt để trích xuất tasks, resources, relations, consumptions.
    Ví dụ định dạng file:
        task id=1 name=TaskA duration=5
        resource id=1 name=Worker capacity=3
        relation task_id_1=1 task_id_2=2
        consumption task_id=1 resource_id=1 amount=2
    """
    tasks = []
    resources = []
    relations = []
    consumptions = []
    
    with open(file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
                
            if line.startswith('task'):
                parts = line.split()
                task = {
                    'id': int(parts[1].split('=')[1]),
                    'name': parts[2].split('=')[1],
                    'duration': int(parts[3].split('=')[1])
                }
                tasks.append(task)
            elif line.startswith('resource'):
                parts = line.split()
                resource = {
                    'id': int(parts[1].split('=')[1]),
                    'name': parts[2].split('=')[1],
                    'capacity': int(parts[3].split('=')[1])
                }
                resources.append(resource)
            elif line.startswith('relation'):
                parts = line.split()
                relation = {
                    'task_id_1': int(parts[1].split('=')[1]),
                    'task_id_2': int(parts[2].split('=')[1])
                }
                relations.append(relation)
            elif line.startswith('consumption'):
                parts = line.split()
                consumption = {
                    'task_id': int(parts[1].split('=')[1]),
                    'resource_id': int(parts[2].split('=')[1]),
                    'amount': int(parts[3].split('=')[1])
                }
                consumptions.append(consumption)
    
    return tasks, resources, relations, consumptions

# ----------------- Hàm xử lý toàn bộ dataset -----------------

def process_dataset(dataset_root, output_file):
    """
    Duyệt qua tất cả các thư mục con trong dataset, đọc các file .smt,
    gbt cho từng instance và xuất kết quả ra file Excel.
    """
    current_task_id = 1  
    
    for root, dirs, files in os.walk(dataset_root):
        for file in files:
            if file.endswith('.smt'):
                file_path = os.path.join(root, file)
                print(f"Processing file: {file_path}")
                try:
                    tasks, resources, relations, consumptions = parse_smt_file(file_path)
                    
                    max_time = sum(task['duration'] for task in tasks) * 2
                    
                    # gbt
                    model, vf, variables, clauses, status = solve_rcpsp(
                        max_time=max_time,
                        tasks=tasks,
                        relations=relations,
                        consumptions=consumptions,
                        resources=resources
                    )
                    
                    # Giải mã kết quả nếu tìm được lời giải
                    schedule = []
                    if status == "SAT" and model:
                        schedule = decode_solution(tasks, model, vf, consumptions)
                    
                    current_task_id = export_schedule_to_xlsx(
                        schedule=schedule,
                        variables=variables,
                        clauses=clauses,
                        status=status,
                        tasks=tasks,
                        resources=resources,
                        relations=relations,
                        start_task_id=current_task_id,
                        output_file=output_file,
                        append=os.path.exists(output_file)
                    )
                    
                except Exception as e:
                    print(f"Lỗi khi xử lý file {file_path}: {str(e)}")
                    continue

# ----------------- Main -----------------

if __name__ == "__main__":
    dataset_path = "SAT\dataset" # data path
    output_excel = "rcpsp_results.xlsx"
    
    if os.path.exists(output_excel):
        os.remove(output_excel)
        
    process_dataset(dataset_path, output_excel)